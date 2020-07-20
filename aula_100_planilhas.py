# Planilhas do Excel em Python
# Documentação-> https://openpyxl.readthedocs.io/en/stable/
# pip install openpyxl

import openpyxl

# Para carregar a planilha, passamos o caminho no método load_workbook
# Como está no mesmo diretório desse arquivo, basta passar o nome dela
arquivo = openpyxl.load_workbook('aula_100_planilha.xlsx')

# Para ver os nomes das planilhas salvas no arquivo
nome_planilhas = arquivo.sheetnames
planilha1 = arquivo['Pedidos']

# Pegar o valor de determinada célula
print(planilha1['b4'].value)

# Pegar os valores de uma coluna
print('Coluna B----------------------------------')
for cell in planilha1['b']:
    print(cell.value)

# Pegar os valores de uma linha
print('Linha 1-----------------------------------')
for cell in planilha1['1']:
    print(cell.value)

# Pegar os valores em um intervalo
print('Intervalo a1:c2---------------------------')
for linha in planilha1['a1:c2']:
    for cell in linha:
        print(cell.value)

# Percorrer a planilha inteira
print('Planilha completa-------------------------')
for linha in planilha1:
    print(linha[0].value, linha[1].value, linha[2].value)

# Alterar/Inserir dado em uma planilha
planilha1['d1'].value = 'Quantidade'
for i in range(2, 5):
    planilha1[f'd{i}'] = i * 3

# Ao fazer isso não alteramos o arquivo orginal, então precisamos salvar as
# alterações em uma nova planilha
arquivo.save('aula_100_nova_planilha.xlsx')


# Criar um arquivo com duas planilhas
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)
planilha.save('aula_100_planilha_criada.xlsx')
